import os
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from dateutil.relativedelta import relativedelta
from datetime import datetime
import requests
import re
import logging
import time
import imghdr


# configuring the logger
logger = logging.getLogger(__name__)
logging.basicConfig(filename='output/robot.log', level=logging.INFO)


class Bot:
    """
    A bot class for scraping news websites.
    """

    def __init__(self, website, phrase, topic, page):
        logger.info("starting robot")

        self.website = website
        self.phrase = phrase
        self.topic = topic
        self.page = int(page)

        # start the web driver
        try:
            logger.info("starting web driver for robot")
            options = FirefoxOptions()
            options.add_argument("--headless")
            self.driver = webdriver.Firefox(options=options)
            # set implicit wait time to 30 seconds
            self.driver.implicitly_wait(30)
        except Exception:
            logger.exception("error starting web driver")
            self.driver.quit()

        # check if workbook exist or create new one.
        try:
            logger.info("looking for excel workbook")
            wb = load_workbook("output/news.xlsx")
        except FileNotFoundError:
            logger.info("Workbook not found, creating new one")
            wb = Workbook()
            ws = wb.active

            # Add column header to worksheet
            header = ['title', 'date', 'description', 'image_name',
                      'title_phrase_count', 'description_phrase_count', 'title_contain_money']
            ws.append(header)

            # Save and close the workbook
            wb.save('output/news.xlsx')
            wb.close()
        except Exception:
            logger.exception("error loading/creating workbook")
            self.driver.quit()

    def transverse(self):
        """
        bot method for navigating web pages
        """
        try:
            logger.info("searching for yahoo latest news")

            driver = self.driver
            driver.get(self.website)

            # wait for browser to open
            WebDriverWait(driver, 30).until(EC.number_of_windows_to_be(1))
            print("firefox web browser opened by driver\n")

            # get the current tab as previous tab and  get the current url
            previous_tab = driver.current_window_handle
            current_url = driver.current_url
            print(f"bot url location at : {current_url}\n")

            if "consent" in current_url:
                # scroll content to show reject button
                scroll_btn = driver.find_element(By.ID, "scroll-down-btn")
                scroll_btn.click()

                # find the reject-all button and click it
                reject_btn = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "button.reject-all"))
                    )
                reject_btn.click()

                # wait for page navigation back to yahoo home 
                WebDriverWait(driver, 30).until(EC.url_changes(current_url))
                current_url = driver.current_url
                print(f"bot url location at : {current_url}\n")              
            
            # find the search input and enter the search phrase
            search_input = driver.find_element(By.ID, "ybar-sbq")
            search_input.send_keys(self.phrase)

            # find, click the search button and navigate to search result
            search_btn = driver.find_element(By.ID, "ybar-search")
            search_btn.click()

            # Wait for the new window to open
            WebDriverWait(driver, 30).until(EC.number_of_windows_to_be(2))

            # Switch to the new window
            open_tabs = driver.window_handles
            for tab in open_tabs:
                if tab is not previous_tab:
                    driver.switch_to.window(tab)

            # wait for page navigation and get the new page title
            WebDriverWait(driver, 30).until(EC.url_changes(current_url))
            current_url = driver.current_url
            print(f"bot url location at : {current_url}\n")

            # click the news button and navigate to news category
            news_category = driver.find_element(
                By.CSS_SELECTOR, "a[href*='https://news.search.yahoo.com/search']")
            news_category.click()

            #  wait for page navigation
            WebDriverWait(driver, 30).until(EC.url_changes(current_url))
            current_url = driver.current_url
            print(f"bot url location at : {current_url}\n")

            # find and scrape latest news in first 3 pages
            page_no = 0
            for x in range(self.page):
                page_no += 1
                logger.info(
                    f"scraping yahoo news on search result page({page_no}) ...")
                self.scrape(driver.find_elements(By.CSS_SELECTOR,
                                                 ".searchCenterMiddle>li"))
                # find and click the next button for page navigations
                next_btn = driver.find_element(By.CSS_SELECTOR, "a.next")
                next_btn.click()

                # wait for page navigation
                WebDriverWait(driver, 30).until(EC.url_changes(current_url))
                current_url = driver.current_url
                print(f"bot url location at : {current_url}\n")

            # end the bot process
            self.finish()
        except Exception:
            logger.exception("error navigating yahoo website")
            self.driver.quit()

    def scrape(self, news):
        """
        bot method for scraping news datas
        """
        try:
            data_rows = []
            for new in news:
                # get new properties
                title_raw = new.find_element(By.CSS_SELECTOR, ".s-title").text
                date_raw = new.find_element(By.CSS_SELECTOR, ".s-time").text
                description_raw = new.find_element(
                    By.CSS_SELECTOR, ".s-desc").text

                # clean the raw properties
                title = self.clean(title_raw)
                date = self.parse_date(self.clean(date_raw))
                description = self.clean(description_raw)

                # search phrase counts in title and decriptions
                description_phrase_count = description.lower().count(self.phrase.lower())
                title_phrase_count = title.lower().count(self.phrase.lower())

                # get the news image and download it
                image_name = ""
                try:
                    image = new.find_element(By.CSS_SELECTOR, ".s-img")
                    image_name = image.get_attribute("alt")
                    image_url = image.get_attribute("src")
                    image_name = self.download_image(image_name, image_url)
                except Exception:
                    logger.warning("no image found for news")
                    image_name = "no image found"

                # check if title contain money
                title_contain_money = self.contains_monetary_value(
                    title, description)

                # append row to data
                data_rows.append([title, date, description, image_name,
                            title_phrase_count, description_phrase_count, title_contain_money])

            self.data_entry(data_rows)
        except Exception:
            logger.exception("error scraping yahoo news")
            self.driver.quit()

    def data_entry(self, data_rows):
        """
        bot method for excel data entry processes
        """
        logger.info("writing news data to an excel file ...")
        try:
            logger.info("writing news data to an excel file")
            # load news workbook
            wb = load_workbook("output/news.xlsx")
            ws = wb.active

            # append data rows to active worksheet
            for data in data_rows:
                ws.append(data)

            # save and close workbook
            wb.save('output/news.xlsx')
            wb.close()

        except Exception:
            logger.exception("error writing news data")
            self.driver.quit()

    def download_image(self, image_name, image_url):
        """
        bot method for downloading new images
        """
        logger.info("downloading news image ...")
        # get the image
        response = requests.get(image_url)

        image_extension = imghdr.what(None, h=response.content)
        if not image_extension or image_extension == "jpeg":
            # Default to jpg if image type cannot be determined
            image_extension = "jpg"

        # Construct the image filename
        image_name = f"{image_name}.{image_extension}"

        # build and create path for the image
        image_path = os.path.join('output', image_name)
        os.makedirs(os.path.dirname(image_path), exist_ok=True)

        # save the image
        with open(image_path, "wb") as f:
            f.write(response.content)

        # return image name
        return image_name

    def parse_date(self, date_str):
        """
        bot method for formating news data
        """
        now = datetime.now()
        if "hour" in date_str:
            hours_ago = int(date_str.split()[0])
            date = now - relativedelta(hours=hours_ago)
        elif "day" in date_str:
            days_ago = int(date_str.split()[0])
            date = now - relativedelta(days=days_ago)
        elif "week" in date_str:
            weeks_ago = int(date_str.split()[0])
            date = now - relativedelta(weeks=weeks_ago)
        elif "month" in date_str:
            months_ago = int(date_str.split()[0])
            date = now - relativedelta(months=months_ago)
        elif "year" in date_str:
            years_ago = int(date_str.split()[0])
            date = now - relativedelta(years=years_ago)
        else:
            date = now

        return date.strftime("%Y-%m-%d %H:%M:%S")

    def contains_monetary_value(self, title, description):
        """
        bot method for finding money patterns in title and description
        """
        # Define a regular expression pattern to match monetary values
        monetary_pattern = r'\$[\d,]+(?:\.\d+)?(?:\s*(?:dollars|USD))?'

        # Search for the pattern in the title
        match = re.search(monetary_pattern, title)

        # if match is none, search for pattern in description
        if not match:
            match = re.search(monetary_pattern, description)

        # Return True if a match is found, False otherwise
        return match is not None

    def clean(self, text):
        """
        bot method for cleaning scraped data
        """
        # Remove every '.' character
        text = text.replace('.', '')

        # Remove every '·' character
        text = text.replace('·', '')

        # Remove double white space
        text = ' '.join(text.split())

        # Remove trailing and leading white space
        clean_text = text.strip()

        return clean_text

    def finish(self):
        """
        bot method that close the driver and end processes
        """
        self.driver.quit()
        logger.info("bot automation process completed successfully")
        logger.info("bot stopped and exited successfully")
